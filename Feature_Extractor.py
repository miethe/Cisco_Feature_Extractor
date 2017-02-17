from Tkinter import Tk
from tkFileDialog import askopenfilename
from openpyxl import Workbook
from openpyxl import load_workbook

Tk().withdraw()
print "Open Feature List"
file = askopenfilename()

print "Open Configs"
configs = askopenfilename()

ip_wb = load_workbook(file)
ip_ws = ip_wb.active

row_int = 1
row_max = ip_ws.max_row
col_max = ip_ws.max_column

configs = open(configs, 'r')

text_file = open("Features.txt", "w")

for line in configs:
	new_row = 1
	while new_row < row_max:
		if line and ip_ws.cell(row=new_row,column=2).value and ip_ws.cell(row=new_row,column=2).value in line and not ip_ws.cell(row=new_row,column=4).value:
			if ip_ws.cell(row=new_row,column=3).value and "Default enabled" in ip_ws.cell(row=new_row,column=3).value and "no" not in line: #default feature and not disabled
			 	text_file.write(ip_ws.cell(row=new_row,column=1).value+"\n")
			 	ip_ws.cell(row=new_row,column=4).value = 1
			elif not ip_ws.cell(row=new_row,column=3).value:
				text_file.write(ip_ws.cell(row=new_row,column=1).value+"\n")
				ip_ws.cell(row=new_row,column=4).value = 1
			elif "Default enabled" not in ip_ws.cell(row=new_row,column=3).value: #feature in configs
				text_file.write(ip_ws.cell(row=new_row,column=1).value+"\n")
				ip_ws.cell(row=new_row,column=4).value = 1
			
		new_row+=1

text_file.close()
